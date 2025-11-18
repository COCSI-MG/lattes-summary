#!/usr/bin/env python
# encoding: utf-8

import streamlit as st
import pandas as pd
import os
import json
import re
import io
import sys
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Adiciona o diretório 'src' ao path para importar módulos do projeto
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', 'src')))
from db.connect import connect, get_db_path
from services.atividade import AtividadeService
from services.categoria import CategoriaService
from pages.components.tables import render_tabela_atividades_pontos, render_relacao_categoria_atividade
from lattes.indice_r import Trabalho


def load_atividade_pontos_from_db() -> dict:
    """
    Carrega os dados da tabela atividade_ponto e retorna como dicionário.
    
    Returns:
        Dicionário com atividade -> pontos
    """
    try:
        conn = connect()
        cursor = conn.cursor()
        
        cursor.execute("SELECT atividade, pontos FROM atividade_ponto")
        rows = cursor.fetchall()
        
        atividade_pontos = {row['atividade']: row['pontos'] for row in rows}
        
        conn.close()
        return atividade_pontos
    except Exception as e:
        st.error(f"Erro ao carregar atividades do banco: {e}")
        return {}


def load_categoria_atividade_from_db() -> dict:
    """
    Carrega as relações entre categorias e atividades do banco de dados.
    Categorias sem atividade associada retornam None.
    
    Returns:
        Dicionário com categoria -> atividade (ou None se não houver relação)
    """
    try:
        conn = connect()
        cursor = conn.cursor()
        
        cursor.execute("SELECT id, categoria FROM categoria_lattes")
        categorias = {row['id']: row['categoria'] for row in cursor.fetchall()}
        
        cursor.execute("""
            SELECT 
                c.categoria as categoria_nome,
                a.atividade as atividade_nome
            FROM categoria_atividade ca
            JOIN categoria_lattes c ON ca.categoria_id = c.id
            JOIN atividade_ponto a ON ca.atividade_id = a.id
        """)
        relacoes = cursor.fetchall()
        
        categoria_para_atividade = {cat_nome: None for cat_nome in categorias.values()}
        
        for row in relacoes:
            categoria_para_atividade[row['categoria_nome']] = row['atividade_nome']
        
        conn.close()
        return categoria_para_atividade
    except Exception as e:
        st.error(f"Erro ao carregar categorias do banco: {e}")
        return {}


def initialize_session_state():
    if 'indice_r_atividade_pontos' not in st.session_state:
        atividade_pontos_db = load_atividade_pontos_from_db()
        st.session_state['indice_r_atividade_pontos'] = atividade_pontos_db

    if 'indice_r_df' not in st.session_state:
        st.session_state['indice_r_df'] = pd.DataFrame(
            columns=['filtro', 'atividade', 'nome do trabalho', 'autores', 'ano', 'pontos']
        )

    if 'indice_r_categoria_para_atividade' not in st.session_state:
        categoria_atividade_db = load_categoria_atividade_from_db()
        st.session_state['indice_r_categoria_para_atividade'] = categoria_atividade_db


def compute_indice_r(producoes_df: pd.DataFrame, atividade_pontos: dict) -> int:
    """
    Calcula o Índice R como a soma ponderada das atividades presentes no DataFrame.
    Se a coluna 'pontos' estiver preenchida, ela é utilizada diretamente; caso contrário,
    utiliza-se o mapa atividade -> pontos.
    """
    if producoes_df is None or producoes_df.empty:
        return 0

    if 'pontos' in producoes_df.columns:
        pontos_series = pd.to_numeric(producoes_df['pontos'], errors='coerce').fillna(0)
        if pontos_series.sum() > 0:
            return int(pontos_series.sum())

    if 'atividade' in producoes_df.columns:
        return int(producoes_df['atividade'].map(atividade_pontos).fillna(0).sum())

    return 0


def extract_json_from_html(html_content: str, nome_arquivo: str = '', categoria: str = ''):
    """
    Extrai os dados JSON do HTML e retorna uma lista de objetos Trabalho.
    
    Args:
        html_content: Conteúdo HTML a ser processado
        nome_arquivo: Nome do arquivo HTML de origem
        categoria: Categoria do trabalho (ex: "Artigos completos em periódicos")
        
    Returns:
        Lista de objetos Trabalho com as informações extraídas
    """
    try:
        match = re.search(r'const DATA = (\[.*?\]);', html_content, re.DOTALL)
        if match:
            json_str = match.group(1)
            data = json.loads(json_str)
            
            trabalhos = []
            for item in data:
                titulo = get_first_value(
                    item,
                    ['titulo', 'título', 'title', 'nome', 'nome do trabalho', 'nome_do_trabalho', 'titulo_do_trabalho', 'descricao', 'descrição', 'description', 'apresentacao', 'evento'],
                    default=''
                )
                autores = get_first_value(
                    item, 
                    ['autores', 'Autores', 'authors', 'orientado_a', 'orientado', 'orientador'], 
                    default=''
                )
                ano = get_first_value(
                    item, 
                    ['ano', 'Ano', 'year', 'Year'], 
                    default=''
                )
                doi = get_first_value(
                    item,
                    ['doi', 'DOI', 'Doi'],
                    default=''
                )
                issn = get_first_value(
                    item,
                    ['issn', 'ISSN', 'Issn'],
                    default=''
                )
                isbn = get_first_value(
                    item,
                    ['isbn', 'ISBN', 'Isbn'],
                    default=''
                )
                
                texto_completo = json.dumps(item, ensure_ascii=False)
                
                campos_conhecidos = {
                    'titulo', 'título', 'title', 'nome', 'nome do trabalho', 'nome_do_trabalho', 'titulo_do_trabalho',
                    'descricao', 'descrição', 'description', 'evento', 'apresentacao',
                    'autores', 'Autores', 'authors', 'orientado_a', 'orientado', 'orientador',
                    'ano', 'Ano', 'year', 'Year',
                    'doi', 'DOI', 'Doi',
                    'issn', 'ISSN', 'Issn',
                    'isbn', 'ISBN', 'Isbn'
                }
                outros = {k: v for k, v in item.items() if k not in campos_conhecidos}
                
                trabalho = Trabalho(
                    autores=autores,
                    titulo=titulo,
                    categoria=categoria,
                    atividade='',  # Será determinado posteriormente pelas regras
                    pontos=None,   # Será calculado posteriormente
                    ano=str(ano) if ano else '',
                    doi=doi,
                    issn=issn,
                    isbn=isbn,
                    outros=outros,
                    texto_completo=texto_completo,
                    nome_arquivo=nome_arquivo
                )
                
                trabalhos.append(trabalho)
            
            return trabalhos
        return []
    except Exception as e:
        return []


def get_first_value(d: dict, keys: list, default=None):
    for k in keys:
        if k in d and d[k] not in (None, ''):
            return d[k]
    return default


def build_categorias(config_options: dict) -> list:
    """
    Define os arquivos e descrições a carregar (evita páginas de "Total").
    Retorna lista de tuplas: (arquivo, descricao, habilitado)
    """
    cfg = config_options or {}
    categorias = [
        # Produções Bibliográficas
        ('PB0-0.html', 'Artigos completos em periódicos', cfg.get('rel_artigo_periodico', True)),
        ('PB1-0.html', 'Livros publicados', cfg.get('rel_livro', True)),
        ('PB2-0.html', 'Capítulos de livros', cfg.get('rel_capitulo_livro', True)),
        ('PB3-0.html', 'Textos em jornais', cfg.get('rel_jornal', True)),
        ('PB4-0.html', 'Trabalhos completos em congressos', cfg.get('rel_trabalho_completo', True)),
        ('PB5-0.html', 'Resumos expandidos', cfg.get('rel_resumo_expandido', True)),
        ('PB6-0.html', 'Resumos em congressos', cfg.get('rel_resumo', True)),
        ('PB7-0.html', 'Artigos aceitos', cfg.get('rel_artigo_aceito', True)),
        ('PB8-0.html', 'Apresentações de trabalho', cfg.get('rel_apresentacao', True)),
        ('PB9-0.html', 'Outros tipos', cfg.get('rel_outro_biblio', True)),

        # Produções Técnicas
        ('PT0-0.html', 'Software com registro', cfg.get('rel_soft_registro', True)),
        ('PT1-0.html', 'Software sem registro', cfg.get('rel_soft_sem_registro', True)),
        ('PT2-0.html', 'Produtos tecnológicos', cfg.get('rel_produto_tec', True)),
        ('PT3-0.html', 'Processos ou técnicas', cfg.get('rel_processo', True)),
        ('PT4-0.html', 'Trabalhos técnicos', cfg.get('rel_trabalho_tec', True)),
        ('PT5-0.html', 'Outros tipos (técnicas)', cfg.get('rel_outro_tec', True)),
        ('PT6-0.html', 'Entrevistas e comentários', cfg.get('rel_entrevista', True)),

        # Produções Artísticas
        ('PA-0.html', 'Total de produções artísticas', cfg.get('rel_prod_artistica', True)),

        # Orientações em Andamento
        ('OA0-0.html', 'Pós-doutorado (andamento)', cfg.get('orient_and_pos_doc', True)),
        ('OA1-0.html', 'Doutorado (andamento)', cfg.get('orient_and_doc', True)),
        ('OA2-0.html', 'Mestrado (andamento)', cfg.get('orient_and_mest', True)),
        ('OA3-0.html', 'Especialização (andamento)', cfg.get('orient_and_esp', True)),
        ('OA4-0.html', 'TCC (andamento)', cfg.get('orient_and_tcc', True)),
        ('OA5-0.html', 'Iniciação científica (andamento)', cfg.get('orient_and_ic', True)),
        ('OA6-0.html', 'Outros tipos (andamento)', cfg.get('orient_and_outro', True)),

        # Orientações Concluídas
        ('OC0-0.html', 'Pós-doutorado (concluída)', cfg.get('orient_conc_pos_doc', True)),
        ('OC1-0.html', 'Doutorado (concluída)', cfg.get('orient_conc_doc', True)),
        ('OC2-0.html', 'Mestrado (concluída)', cfg.get('orient_conc_mest', True)),
        ('OC3-0.html', 'Especialização (concluída)', cfg.get('orient_conc_esp', True)),
        ('OC4-0.html', 'TCC (concluída)', cfg.get('orient_conc_tcc', True)),
        ('OC5-0.html', 'Iniciação científica (concluída)', cfg.get('orient_conc_ic', True)),
        ('OC6-0.html', 'Outros tipos (concluída)', cfg.get('orient_conc_outro', True)),

        # Outros
        ('Pj-0.html', 'Projetos de pesquisa', cfg.get('rel_projeto', True)),
        ('Pm-0.html', 'Prêmios e títulos', cfg.get('rel_premio', True)),
        ('Ep-0.html', 'Participação em eventos', cfg.get('rel_part_evento', True)),
        ('Eo-0.html', 'Organização de eventos', cfg.get('rel_org_evento', True)),
    ]
    return categorias


def carregar_dados_para_dataframe(output_dir: str) -> pd.DataFrame:
    """
    Carrega os dados dos arquivos HTML e retorna um DataFrame com os trabalhos.
    Utiliza a estrutura de objetos Trabalho do método extract_json_from_html.
    
    Args:
        output_dir: Diretório contendo os arquivos HTML
        
    Returns:
        DataFrame com as colunas: id, filtro, atividade, nome do trabalho, autores, ano, pontos
    """
    atividade_pontos = st.session_state['indice_r_atividade_pontos']
    cat_to_atividade = st.session_state['indice_r_categoria_para_atividade']
    config_options = st.session_state.get('config_options', {})
    categorias = build_categorias(config_options)

    linhas = []
    next_id = 1

    for arquivo, descricao, habilitado in categorias:
        if not habilitado:
            continue

        if descricao.lower().startswith('total de'):
            continue

        file_path = os.path.join(output_dir, arquivo)
        if not os.path.exists(file_path):
            continue

        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                html_content = f.read()
            trabalhos = extract_json_from_html(html_content, nome_arquivo=arquivo, categoria=descricao)
        except Exception:
            trabalhos = []

        if not trabalhos:
            continue

        for trabalho in trabalhos:
            atividade = trabalho.atividade if trabalho.atividade else cat_to_atividade.get(descricao)
            
            pontos = trabalho.pontos if trabalho.pontos is not None else (atividade_pontos.get(atividade) if atividade else None)

            linhas.append({
                'id': next_id,
                'filtro': trabalho.categoria,  # usa a categoria do objeto Trabalho
                'atividade': atividade,
                'nome do trabalho': trabalho.titulo if trabalho.titulo else '—',
                'autores': trabalho.autores if trabalho.autores else '—',
                'ano': trabalho.ano,
                'pontos': pontos,
            })
            next_id += 1

    if not linhas:
        return pd.DataFrame(columns=['id', 'filtro', 'atividade', 'nome do trabalho', 'autores', 'ano', 'pontos'])

    df = pd.DataFrame(linhas)
    return df


def main():
    st.title('📈 Índice R')
    st.markdown('---')

    initialize_session_state()

    atividade_pontos = st.session_state['indice_r_atividade_pontos']
    cat_to_atividade = st.session_state['indice_r_categoria_para_atividade']

    output_dir = st.session_state.get('output_dir', 'tmp')
    st.caption(f"Diretório de dados: `{output_dir}`")

    col_load1, col_load2 = st.columns(2)
    with col_load1:
        if st.button('🔄 Recarregar dados a partir dos arquivos', use_container_width=True):
            df_loaded = carregar_dados_para_dataframe(output_dir)
            st.session_state['indice_r_df'] = df_loaded
            st.rerun()
    with col_load2:
        if st.button('🧹 Limpar dados do Índice R', use_container_width=True):
            st.session_state['indice_r_df'] = pd.DataFrame(columns=['id', 'filtro', 'atividade', 'nome do trabalho', 'autores', 'ano', 'pontos'])
            st.rerun()

    if st.session_state['indice_r_df'].empty and os.path.isdir(output_dir):
        st.session_state['indice_r_df'] = carregar_dados_para_dataframe(output_dir)

    producoes_df = st.session_state['indice_r_df']

    indice_r = compute_indice_r(producoes_df, atividade_pontos)

    col1, col2 = st.columns(2)
    with col1:
        st.metric(label='Índice R', value=f'{indice_r}')
    with col2:
        total_itens = 0 if producoes_df is None else len(producoes_df)
        st.metric(label='Total de Itens', value=f'{total_itens}')

    st.markdown('---')

    render_tabela_atividades_pontos(load_atividade_pontos_from_db)
    render_relacao_categoria_atividade(load_categoria_atividade_from_db)

    if not producoes_df.empty:
        df_sem = producoes_df[producoes_df['atividade'].isna()].copy()
        df_com = producoes_df[producoes_df['atividade'].notna()].copy()

        st.subheader('Categorias sem atividade definida')
        st.caption('Atribua uma atividade para cada item abaixo.')
        if df_sem.empty:
            st.success('Nenhuma categoria pendente de mapeamento. ✅')
            edited_sem = df_sem
        else:
            atividade_options = list(atividade_pontos.keys())
            edited_sem = st.data_editor(
                df_sem,
                key='editor_sem_atividade',
                hide_index=True,
                use_container_width=True,
                height=300,
                column_config={
                    'id': st.column_config.NumberColumn('id', disabled=True, width='small'),
                    'filtro': st.column_config.TextColumn('filtro', disabled=True),
                    'atividade': st.column_config.SelectboxColumn('atividade', options=atividade_options, required=False, help='Selecione a atividade'),
                    'nome do trabalho': st.column_config.TextColumn('nome do trabalho', disabled=True),
                    'autores': st.column_config.TextColumn('autores', disabled=True),
                    'ano': st.column_config.TextColumn('ano', disabled=True),
                    'pontos': st.column_config.NumberColumn('pontos', disabled=True),
                },
            )
        

        st.markdown('---')

        st.subheader('Itens com atividade (você pode alterar se necessário)')
        atividade_options = list(atividade_pontos.keys())
        edited_com = st.data_editor(
            df_com,
            key='editor_com_atividade',
            hide_index=True,
            use_container_width=True,
            height=360,
            column_config={
                'id': st.column_config.NumberColumn('id', disabled=True, width='small'),
                'filtro': st.column_config.TextColumn('filtro', disabled=True),
                'atividade': st.column_config.SelectboxColumn('atividade', options=atividade_options, required=False),
                'nome do trabalho': st.column_config.TextColumn('nome do trabalho', disabled=True),
                'autores': st.column_config.TextColumn('autores', disabled=True),
                'ano': st.column_config.TextColumn('ano', disabled=True),
                'pontos': st.column_config.NumberColumn('pontos', disabled=True),
            },
        )

        df_atualizado = producoes_df.copy()
        if not edited_sem.empty:
            for _, row in edited_sem.iterrows():
                rid = row['id']
                nova_atividade = row['atividade'] if row['atividade'] != '' else None
                df_atualizado.loc[df_atualizado['id'] == rid, 'atividade'] = nova_atividade
        if not edited_com.empty:
            for _, row in edited_com.iterrows():
                rid = row['id']
                nova_atividade = row['atividade'] if row['atividade'] != '' else None
                df_atualizado.loc[df_atualizado['id'] == rid, 'atividade'] = nova_atividade
        df_atualizado['pontos'] = df_atualizado['atividade'].map(atividade_pontos).fillna(0)

        st.markdown('---')

        template_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'template', 'tabela_indice_R.xlsx'))
        if not os.path.isfile(template_path):
            st.error('Template XLSX não encontrado em src/template/tabela_indice_R.xlsx')
        else:
            try:
                wb = load_workbook(template_path)
                if 'base' in wb.sheetnames:
                    ws = wb['base']
                    wb.remove(ws)
                    ws = wb.create_sheet('base', 0)
                else:
                    ws = wb.create_sheet('base', 0)

                for row in dataframe_to_rows(df_atualizado, index=False, header=True):
                    ws.append(row)

                buffer = io.BytesIO()
                wb.save(buffer)
                buffer.seek(0)

                st.download_button(
                    label='📤 Exportar XLSX (template)',
                    data=buffer,
                    file_name='indice_r.xlsx',
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    use_container_width=True
                )
            except Exception as e:
                st.error(f'Falha ao gerar XLSX: {e}')

        if st.button('💾 Aplicar alterações em atividades', type='primary', use_container_width=True):
            df_atual = st.session_state['indice_r_df']
            if not df_sem.empty:
                for _, row in edited_sem.iterrows():
                    rid = row['id']
                    nova_atividade = row['atividade'] if row['atividade'] != '' else None
                    df_atual.loc[df_atual['id'] == rid, 'atividade'] = nova_atividade
            for _, row in edited_com.iterrows():
                rid = row['id']
                nova_atividade = row['atividade'] if row['atividade'] != '' else None
                df_atual.loc[df_atual['id'] == rid, 'atividade'] = nova_atividade

            df_atual['pontos'] = df_atual['atividade'].map(atividade_pontos).fillna(0)

            st.session_state['indice_r_df'] = df_atual
            st.rerun()


if __name__ == '__main__':
    main()


